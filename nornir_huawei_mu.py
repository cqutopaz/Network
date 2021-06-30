#!/usr/bin/python3
#coding=utf-8

from nornir import InitNornir
from nornir.core.filter import F
from nornir_utils.plugins.functions import print_result,print_title
from nornir_utils.plugins.tasks.files import write_file
from nornir_netmiko import netmiko_send_command
from nornir.core.task import Result,Task
import os
import time
import ipdb
import json
import pandas as pd
import datetime
import glob							#用于筛选获取文件夹内文件信息
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Border,Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

start_time = time.perf_counter()
print("脚本执行开始时间："+time.strftime('%X'))
nr = InitNornir(config_file = 'config.yaml')
#用F方法过滤出相关设备，可以使用'|'这个‘或’标志来同时筛选
#定义ntc-textfsm模板源到我自定义的路径及其templates
os.environ["NET_TEXTFSM"] = "/home/cqutopaz/custom_ntc_templates/templates"
#定义筛选对象的变量，即定义筛选后的设备
#sws = ['sw1','sw2','sw3','sw4']
#sw = nr.filter(filter_func = lambda host: host.name in sws)
huawei_sw = nr.filter(F(platform = 'huawei'))

#检查要创建的存放目录是否存在
pwd = os.getcwd()
filename = pwd+'/output_excel'+datetime.datetime.now().strftime('%Y%m%d')
if not os.path.exists(filename):
	os.mkdir(filename)
result_file='Inspection_'+datetime.datetime.now().strftime('%Y-%m-%d')+'.xlsx'

#定义一个任务开始描述的对象
def say_sth(task:Task,sth:str='') -> Result:
	'''让设备打个招呼哈'''
	words = f"Hello!I'm {sth}.My address is {task.host.hostname}"
	return Result(
		host = task.host,
		result = words)
#萝卜橙子西红柿定义要巡检的项目清单（字典）
cmds = {'display version':'version_info',
		'display device':'device_info',
		'display device manufacture-info':'device_serial_info',
		'display cpu-usage':'cpu_info',
		'display transceiver verbose':'transceiver_info'}
#萝卜橙子西红柿定义一个登陆设备获取巡检输出json，并转换到xls的函数
def Inspec(task:Task) -> Result:
	for cmd,dsth in cmds.items():
		results = task.run(netmiko_send_command,command_string = cmd,use_textfsm = True)
#		print('results的类型是：'+str(type(results)))  #results的类型是：<class 'nornir.core.task.MultiResult'>
		output=results.result       #萝卜橙子西红柿
#		print('results.result的类型是：'+str(type(output)))  #results.result的类型是：<class 'list'>
#萝卜橙子西红柿构建第一列数据：交换机名称
		df1 = pd.DataFrame(columns = ['hostname'],data = [[task.host.name]]*len(output))
#萝卜橙子西红柿构建第二列数据：交换机IP信息
		df2 = pd.DataFrame(columns = ['host_ip'],data = [[task.host.hostname]]*len(output))
#萝卜橙子西红柿构建第二块数据：巡检输出数据
		df3 = pd.DataFrame(output)
#萝卜橙子西红柿拼接上面两块数据，按列方向拼接
		df = pd.concat([df1,df2,df3],axis = 1)
		df.to_excel(filename+'/'+f'{task.host.name}'+'_'+dsth+'.xlsx',sheet_name = f'{task.host.name}',index = False)

#定义一个组合task
def huawei_dis_device(task):
	task.run(task = say_sth,sth = 'huawei switch')
	task.run(task = Inspec)
	return Result(host = task.host,result = "已执行dis device命令")

'''下面是数据清洗excel重塑'''
#定义合并数据到一个excel的一个sheet
def combine_device_excel():
	for cmd,dsth in cmds.items():
		sheetname = dsth
	#存放要合并的目标工作簿
		filearray = []
	#根据dsth巡检项目筛选出来合并的目标工作簿，列表
		path = glob.glob(filename+r'/*'+dsth+'.xlsx')
	#排序，使巡检对象按列排序的时候不是杂乱随机的
		filelocation = sorted(path)
	#	print("Filelocation is "+str(filelocation))
		for excel_name in filelocation:
			filearray.append(excel_name)
	#		print(excel_name)
	#	print("Filearray is "+str(filearray))
	#定义目标工作薄res，先读取一个工作簿到定义的目标工作簿内
		res = pd.read_excel(filearray[0])
	#遍历出所有需要合并到一个sheet的工作簿，并定义当前获取到的工作簿为A表，然后concat合并res和A，重复合并
		for i in range(1,len(filearray)):
			A = pd.read_excel(filearray[i])
			res = pd.concat([res,A],ignore_index = True,sort = False)
	#	print(res.index)
		writer = pd.ExcelWriter(filename+'/Ins-'+dsth+'-'+datetime.datetime.now().strftime('%Y-%m-%d')+'.xlsx')
		res.to_excel(writer,sheet_name=sheetname,index = False)
		writer.save()

#定义合并不同excel到一个excel的不同sheet的函数
def combine_excel():
	path = glob.glob(filename+r'/*'+datetime.datetime.now().strftime('%Y-%m-%d')+'.xlsx')
	origin_file_list = sorted(path,key=os.path.getctime)
	writer = pd.ExcelWriter(result_file)
	for ii in origin_file_list:
		sheet_name = ii.split('-')[1]
		df = pd.read_excel(ii)		#读取要合并对象文件的内容
		df.to_excel(writer,sheet_name = sheet_name,index = False)
	writer.save()

##修饰重塑最终获取的excel文件
def remodeling_excel():
	#设置字体格式
	all_font = Font(name = '微软雅黑',size = 12,bold = False,italic=False,underline='none',color='000000')
	ok_result_font = Font(name='微软雅黑', size=12, italic=False, color='00C400', bold=True)
	error_result_font = Font(name='微软雅黑', size=12, italic=False, color='DA0000', bold=True)
	#设置单元格格式为水平居中、垂直居中、不自动换行
	align = Alignment(horizontal='center',vertical='center',wrap_text=False)
	#设置单元格边框，左右上下四个边框类型均为thin
	thin_border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	#用PatternFill()为每个工作表的首行单元格设置背景颜色
	Row1Fill=PatternFill(start_color='C0C0C0',end_color='C0C0C0',fill_type='solid')
	"""合并相同数据到一个单元格"""
#载入要处理的excel文档工作簿
	wb = load_workbook(result_file)
	for ws in wb:
		#设置sheet选项卡背景色
		ws.sheet_properties.tabColor="00FF00"
		re_list = []
#从A2开始抓取数据,获取要合并对象单元格所在列的所有单元格数据，并全部存在re_list这个列表里
		i = 2
		while True:
		    r = ws.cell(i, 1).value
		    if r:
		        re_list.append(r)
		    else:
		        break       #遇到控制跳出循环
		    i += 1
# 判断合并单元格的始末位置
		s = 0
		e = 0
		flag = re_list[0] 
for i in range(len(re_list)):
		    if re_list[i] != flag:
		        flag = re_list[i]
		        e = i - 1
		        if e >= s:
		            ws.merge_cells("A" + str(s + 2) + ":A" + str(e + 2))
		            s = e + 1
		    if i == len(re_list) - 1:
		        e = i
		        ws.merge_cells("A" + str(s + 2) + ":A" + str(e + 2))
# 获取第二列数据
		re2_list = []
		i = 2
		while True:
		    r2 = ws.cell(i, 2).value
		    if r2:
		        re2_list.append(r2)
		    else:
		        break       #遇到控制跳出循环
		    i += 1
		s2 = 0
		e2 = 0
		flag2 = re2_list[0] 
		for i in range(len(re2_list)):
		    if re2_list[i] != flag2:
		        flag2 = re2_list[i]
		        e2 = i - 1
		        if e2 >= s2:
		            ws.merge_cells("B" + str(s2 + 2) + ":B" + str(e2 + 2))
		            s2 = e2 + 1
		    if i == len(re2_list) - 1:
		        e2 = i
		        ws.merge_cells("B" + str(s2 + 2) + ":B" + str(e2 + 2))
#筛选有Normal或Error的单元格并处里
		for Acell in ws['A']:
		    Acell.alignment = align
		for Bcell in ws['B']:
			Bcell.alignment = align
		for column in ws.columns:
		    for cell in column:
		        if cell.value == 'Normal':
		            cell.font = ok_result_font
		        elif cell.value == 'Error':
		            cell.font = error_result_font
		        else:
		            cell.font = all_font
#筛选光功率巡检表里的Rx和Tx
		if ws.title == 'transceiver_info':
			column_rx=ws['I']
			column_tx=ws['J']
			rx_col=[] 
			tx_col=[]
			for rxx in column_rx:
				rx_col.append(rxx.value)
			for i in range(1,len(rx_col)-1):
				if ws['I'+str(i+1)].value < -14:
					ws['I'+str(i+1)].font = error_result_font
				elif ws['I'+str(i+1)].value > 0:
					ws['I'+str(i+1)].font = error_result_font
			for txx in column_tx:
				tx_col.append(txx.value)
			for i in range(1,len(tx_col)-1):
				if ws['J'+str(i+1)].value < -10:
					ws['J'+str(i+1)].font = error_result_font
				elif ws['J'+str(i+1)].value > -1:
					ws['J'+str(i+1)].font = error_result_font
#第一行背景色上色
		row_1=ws[1]
		for cell in row_1:
			cell.fill=Row1Fill
#自适应调整列宽
		dims={}         #存放每列的最大列宽数据
		for row in ws.rows:
			for cell in row:
				cell.border=thin_border
				if cell.value:
					cell_col=get_column_letter(cell.column)
					dims[cell_col]=max(dims.get(cell_col,0),len(str(cell.value)))
		for col,value in dims.items():
			ws.column_dimensions[col].width=value+2
	wb.save(result_file)


def main():
	hw_dis_device_results = huawei_sw.run(task = huawei_dis_device,name = 'A task for excute task_group!')
	combine_device_excel()
	time.sleep(1)
	combine_excel()
	remodeling_excel()

if __name__=='__main__':
	main()
	end_time = time.perf_counter()-start_time
	print("脚本执行结束时间："+time.strftime('%X'))
	print(f'总共耗时{round(end_time,2)}秒')

#ipdb.set_trace()
